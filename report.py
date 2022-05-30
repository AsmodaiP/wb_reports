'''Модуль для работы с файлами отчетов'''

from typing import List

import openpyxl


def find_start_and_end_of_realize(worksheet):
    '''Находит последнюю строку с реализованным товаром по ключевому слову "Итого"'''
    count = 0
    for row in worksheet:
        if row[0].value == '1':
            start = count + 2
        if row[0].value == 'ИТОГО:':
            end = count
            return start, end
        count += 1
    return None


def find_start_and_end_of_refund(worksheet):
    '''Находит первую последнюю строку возвратов'''
    count = 0
    start = 0
    for row in worksheet:
        if row[0].value == 'Возврат реализованного товара текущего периода':
            start = count + 5
        if row[0].value == 'ИТОГО:' and start != 0:
            end = count
            return start, end
        count += 1
    return None


def get_unsorted_relized(name_of_xlsx='0.xlsx'):
    '''Получение неотсортированного списка реализованного товара из файла'''
    workbook = openpyxl.load_workbook(name_of_xlsx)
    worksheet = workbook.active
    # start, end = (find_start_and_end_of_realize(worksheet))
    relized = []
    # for row in worksheet.rows:
    #     print(row)

    base_dict_for_relized = {
        'Наименование': 6,
        'Артикул': 5,
        'Количество': 12,
        'Скидка постоянного покупателя': 16,
        'Вайлдберриз реализовал': 14,
        'Возмещение расходов': 21,
        'Вознаграждение Вайлдберриз без НДС': 22,
        'НДС с вознаграждения Вайлдберриз': 23,
        'К перечислению': 34,
        'Обоснование': 29
    }

    relized_dicts = {}

    for row in worksheet.rows:
        if 'родажа' in row[base_dict_for_relized['Обоснование']].value:
            continue
        tmp = {}
        for key, value in base_dict_for_relized.items():
            tmp[key] = row[value].value
        if tmp['Артикул'] not in relized_dicts.keys():
            relized_dicts[tmp['Артикул']] = tmp
        else:

            relized_dicts[tmp['Артикул']]['Количество']+=int(tmp['Количество'])
            relized_dicts[tmp['Артикул']]['Вайлдберриз реализовал']+=float(tmp['Вайлдберриз реализовал'])
            relized_dicts[tmp['Артикул']]['Вознаграждение Вайлдберриз без НДС']+=float(tmp['Вознаграждение Вайлдберриз без НДС'])
            relized_dicts[tmp['Артикул']]['НДС с вознаграждения Вайлдберриз']+=float(tmp['НДС с вознаграждения Вайлдберриз'])
            relized_dicts[tmp['Артикул']]['К перечислению']+=float(tmp['К перечислению'])

    return(relized_dicts)

    # for row in range(start, end + 1):
    #     tmp = {}
    #     for key, value in base_dict_for_relized.items():
    #         tmp[key] = worksheet[row][value].value
    #     relized_dicts[tmp['Артикул']] = tmp
    # return relized_dicts


def dict_sort(dictionary):
    '''Сортировка словаря'''
    sorted_keys = sorted(dictionary)
    sorted_dict = {}
    for key in sorted_keys:
        sorted_dict[key] = dictionary[key]
    return sorted_dict


def get_sorted_relized(name_of_xlsx='report.xlsx'):
    '''Получение отсортированного словаря с разбиением по префиксу'''
    workbook = openpyxl.load_workbook(name_of_xlsx)
    worksheet = workbook.active
    unsorted = get_unsorted_relized(worksheet)
    sorted_relized = dict_sort(unsorted)
    sorted_dict = {}
    for article, _ in sorted_relized.items():
        if article[:2] not in sorted_dict:
            sorted_dict[article[:2]] = {article: sorted_relized[article]}
        else:
            sorted_dict[article[:2]][article] = sorted_relized[article]
    return sorted_dict


def get_refunds(worksheet):
    '''Получение листа со всеми возвратами из таблицы'''
    start, end = find_start_and_end_of_refund(worksheet)
    refunds = []
    for row in range(start, end + 1):
        refunds.append(worksheet[row])
    return refunds


def group_refunds_by_prefix_and_summ(refunds: List):
    '''Получение листа с префиксами и соответсвующими суммами'''
    refunds_sums = {}
    for refund in refunds:
        prefix = refund[5].value[:2]
        sum_of_refund = float(refund[8].value)
        if prefix not in refunds_sums.keys():
            refunds_sums[prefix] = sum_of_refund
        else:
            refunds_sums[prefix] += sum_of_refund
    return refunds_sums


def get_sum_of_refunds(name_of_xlsx='report.xlsx') -> int:
    '''Возвращает сумму всех возвратов из файла'''
    wb = openpyxl.load_workbook(name_of_xlsx)
    ws = wb.active
    start, end = find_start_and_end_of_refund(ws)
    refunds = []
    s = 0
    for row in range(start, end + 1):
        s += (ws[row][8].value)
    return s


def get_refunds_sums(name_of_xlsx='report.xlsx'):
    wb = openpyxl.load_workbook(name_of_xlsx)
    ws = wb.active
    return group_refunds_by_prefix_and_summ(get_refunds(ws))


if __name__ == '__main__':
    # print(get_sum_of_refunds('0.xlsx'))
    get_unsorted_relized()
